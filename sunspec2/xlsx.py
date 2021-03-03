
"""
    Copyright (C) 2020 SunSpec Alliance

    Permission is hereby granted, free of charge, to any person obtaining a
    Permission is hereby granted, free of charge, to any person obtaining a
    copy of this software and associated documentation files (the "Software"),
    to deal in the Software without restriction, including without limitation
    the rights to use, copy, modify, merge, publish, distribute, sublicense,
    and/or sell copies of the Software, and to permit persons to whom the
    Software is furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included
    in all copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL
    THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING
    FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS
    IN THE SOFTWARE.
"""

import sunspec2.mdef as mdef
import sunspec2.spreadsheet as ss

models_hdr = [('Model', 0),
              ('Label', 30),
              ('Description', 60)]

column_width = {
    ss.ADDRESS_OFFSET_IDX: 0,
    ss.GROUP_OFFSET_IDX: 0,
    ss.NAME_IDX: 25,
    ss.VALUE_IDX: 12,
    ss.COUNT_IDX: 12,
    ss.TYPE_IDX: 12,
    ss.SIZE_IDX: 12,
    ss.SCALE_FACTOR_IDX: 12,
    ss.UNITS_IDX: 12,
    ss.ACCESS_IDX: 12,
    ss.MANDATORY_IDX: 12,
    ss.STATIC_IDX: 12,
    ss.LABEL_IDX: 30,
    ss.DESCRIPTION_IDX: 60,
    ss.NOTES_IDX: 60
}

group_styles = {
    'suns_group_1': {
        'group_color': 'b8cce4', # 184, 204, 228
        'point_color': 'dce6f1', # 220, 230, 241
    },
    'suns_group_2': {
        'group_color': 'd8e4bc', # 216, 228, 188
        'point_color': 'ebf1de', # 235, 241, 222
    },
    'suns_group_3': {
        'group_color': 'ccc0da', # 204, 192, 218
        'point_color': 'e4dfec', # 228, 223, 236
    },
    'suns_group_4': {
        'group_color': 'fcd5b4', # 252, 213, 180
        'point_color': 'fde9d9', # 253, 233, 217
    },
    'suns_group_5': {
        'group_color': 'e6b8b7', # 230, 184, 183
        'point_color': 'f2dcdb', # 242, 220, 219
    }
}

try:
    import openpyxl
    import openpyxl.styles as styles


    class ModelWorkbook(object):
        def __init__(self, filename=None, model_dir=None, license_summary=False, params=None):
            self.wb = None
            self.filename = filename
            self.params = params
            if self.params is None:
                self.params = {}

            if filename is not None:
                self.wb = openpyxl.load_workbook(filename=filename)
            else:
                self.wb = openpyxl.Workbook()

                self.ws_models = self.wb.active
                self.ws_models.title = 'Index'

                thin = styles.Side(border_style=self.params.get('side_border', 'thin'),
                                   color=self.params.get('side_color', '999999'))

                for i in range(1, len(group_styles) + 1):
                    key = 'suns_group_%s' % i
                    name = 'suns_group_entry_%s' % i
                    style = styles.NamedStyle(name=name)
                    color = group_styles[key]['group_color']
                    # self.params.get('group_color', color)
                    style.fill = styles.PatternFill('solid', fgColor=color)
                    style.font = styles.Font()
                    style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(style)

                    name = 'suns_group_text_%s' % i
                    style = styles.NamedStyle(name=name)
                    style.fill = styles.PatternFill('solid', fgColor=color)
                    style.font = styles.Font()
                    style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(style)

                    name = 'suns_point_entry_%s' % i
                    style = styles.NamedStyle(name=name)
                    color = group_styles[key]['point_color']
                    # self.params.get('group_color', color)
                    style.fill = styles.PatternFill('solid', fgColor=color)
                    style.font = styles.Font()
                    style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(style)

                    name = 'suns_point_text_%s' % i
                    style = styles.NamedStyle(name=name)
                    style.fill = styles.PatternFill('solid', fgColor=color)
                    style.font = styles.Font()
                    style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(style)

                if 'suns_hdr' not in self.wb.named_styles:
                    hdr_style = styles.NamedStyle(name='suns_hdr')
                    hdr_style.fill = styles.PatternFill('solid', fgColor=self.params.get('hdr_color', 'dddddd'))
                    hdr_style.font = styles.Font(bold=True)
                    hdr_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    hdr_style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(hdr_style)
                if 'suns_group_entry' not in self.wb.named_styles:
                    model_entry_style = styles.NamedStyle(name='suns_group_entry')
                    model_entry_style.fill = styles.PatternFill('solid',
                                                                fgColor=self.params.get('group_color', 'fff9e5'))
                    model_entry_style.font = styles.Font()
                    model_entry_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    model_entry_style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(model_entry_style)
                if 'suns_group_text' not in self.wb.named_styles:
                    model_text_style = styles.NamedStyle(name='suns_group_text')
                    model_text_style.fill = styles.PatternFill('solid',
                                                               fgColor=self.params.get('group_color', 'fff9e5'))
                    model_text_style.font = styles.Font()
                    model_text_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    model_text_style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(model_text_style)
                if 'suns_point_entry' not in self.wb.named_styles:
                    fixed_entry_style = styles.NamedStyle(name='suns_point_entry')
                    fixed_entry_style.fill = styles.PatternFill('solid',
                                                                fgColor=self.params.get('point_color', 'e6f2ff'))
                    fixed_entry_style.font = styles.Font()
                    fixed_entry_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    fixed_entry_style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(fixed_entry_style)
                if 'suns_point_text' not in self.wb.named_styles:
                    fixed_text_style = styles.NamedStyle(name='suns_point_text')
                    fixed_text_style.fill = styles.PatternFill('solid',
                                                               fgColor=self.params.get('point_color', 'e6f2ff'))
                    fixed_text_style.font = styles.Font()
                    fixed_text_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    fixed_text_style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(fixed_text_style)
                if 'suns_point_variable_entry' not in self.wb.named_styles:
                    fixed_entry_style = styles.NamedStyle(name='suns_point_variable_entry')
                    fixed_entry_style.fill = styles.PatternFill('solid',
                                                                fgColor=self.params.get('point_variable_color', 'ecf9ec'))
                    fixed_entry_style.font = styles.Font()
                    fixed_entry_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    fixed_entry_style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(fixed_entry_style)
                if 'suns_point_variable_text' not in self.wb.named_styles:
                    fixed_text_style = styles.NamedStyle(name='suns_point_variable_text')
                    fixed_text_style.fill = styles.PatternFill('solid',
                                                               fgColor=self.params.get('point_variable_color', 'ecf9ec'))
                    fixed_text_style.font = styles.Font()
                    fixed_text_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    fixed_text_style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(fixed_text_style)
                if 'suns_symbol_entry' not in self.wb.named_styles:
                    repeating_entry_style = styles.NamedStyle(name='suns_symbol_entry')
                    repeating_entry_style.fill =styles.PatternFill('solid',
                                                                   fgColor=self.params.get('symbol_color', 'fafafa'))
                    repeating_entry_style.font = styles.Font()
                    repeating_entry_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    repeating_entry_style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(repeating_entry_style)
                if 'suns_symbol_text' not in self.wb.named_styles:
                    repeating_text_style = styles.NamedStyle(name='suns_symbol_text')
                    repeating_text_style.fill = styles.PatternFill('solid',
                                                                   fgColor=self.params.get('symbol_color', 'fafafa'))
                    repeating_text_style.font = styles.Font()
                    repeating_text_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    repeating_text_style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(repeating_text_style)
                if 'suns_comment' not in self.wb.named_styles:
                    symbol_text_style = styles.NamedStyle(name='suns_comment')
                    symbol_text_style.fill = styles.PatternFill('solid',
                                                                fgColor=self.params.get('comment_color', 'dddddd'))
                                                                # fgColor=self.params.get('symbol_color', 'fffcd9'))
                    symbol_text_style.font = styles.Font()
                    symbol_text_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    symbol_text_style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(symbol_text_style)
                if 'suns_entry' not in self.wb.named_styles:
                    entry_style = styles.NamedStyle(name='suns_entry')
                    entry_style.fill = styles.PatternFill('solid', fgColor='ffffff')
                    entry_style.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                    entry_style.alignment = styles.Alignment(horizontal='center', wrapText=True)
                    self.wb.add_named_style(entry_style)
                if 'suns_text' not in self.wb.named_styles:
                    text_style = styles.NamedStyle(name='suns_text')
                    text_style.font = styles.Font()
                    text_style.alignment = styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(text_style)
                if 'suns_hyper' not in self.wb.named_styles:
                    hyper_style = openpyxl.styles.NamedStyle(name='suns_hyper')
                    hyper_style.font = openpyxl.styles.Font(color='0000ee', underline='single')
                    hyper_style.alignment = openpyxl.styles.Alignment(horizontal='left', wrapText=True)
                    self.wb.add_named_style(hyper_style)

                for i in range(len(models_hdr)):
                    self.set_cell(self.ws_models, 1, i + 1, models_hdr[i][0], 'suns_hdr')
                    if models_hdr[i][1]:
                        self.ws_models.column_dimensions[chr(65 + i)].width = models_hdr[i][1]

        def get_models(self):
            models = []
            if self.wb is not None:
                for m in self.wb.sheetnames:
                    try:
                        mid = int(m)
                        models.append(mid)
                    except:
                        pass
            return models

        def save(self, filename):
            self.wb.save(filename)

        def xlsx_iter_rows(self, ws):
            for row in ws.iter_rows():
                yield [cell.value for cell in row]

        def spreadsheet_from_xlsx(self, mid=None):
            spreadsheet = []
            ws = self.wb[str(mid)]
            for row in self.xlsx_iter_rows(ws):
                # filter out informative offset information from the normative model definition
                if row[ss.TYPE_IDX] and row[ss.TYPE_IDX] != ss.TYPE:
                    row[ss.ADDRESS_OFFSET_IDX] = ''
                    row[ss.GROUP_OFFSET_IDX] = ''
                spreadsheet.append(row)
            return spreadsheet

        def from_xlsx(self, mid=None):
            return ss.from_spreadsheet(self.spreadsheet_from_xlsx(mid))

        def set_cell(self, ws, row, col, value, style=None):
            if self.filename:
                raise ValueError('Workbooks opened with existing file are read only')
            cell = ws.cell(row=row, column=col)
            cell.value = value
            if style:
                cell.style = style
            return cell

        def set_info(self, ws, row, values, style=None):
            self.set_cell(ws, row, ss.LABEL_IDX + 1, values[ss.LABEL_IDX], style=style)
            self.set_cell(ws, row, ss.DESCRIPTION_IDX + 1, values[ss.DESCRIPTION_IDX], style=style)
            if len(values) > ss.NOTES_IDX:
                self.set_cell(ws, row, ss.NOTES_IDX + 1, values[ss.NOTES_IDX], style=style)

        def set_group(self, ws, row, values, level):
            for i in range(len(values)):
                self.set_cell(ws, row, i + 1, '', 'suns_group_entry_%s' % level)
            self.set_cell(ws, row, ss.NAME_IDX + 1, values[ss.NAME_IDX])
            self.set_cell(ws, row, ss.TYPE_IDX + 1, values[ss.TYPE_IDX])
            self.set_cell(ws, row, ss.COUNT_IDX + 1, values[ss.COUNT_IDX])
            self.set_info(ws, row, values, 'suns_group_text_%s' % level)

        def set_point(self, ws, row, values, level):
            entry_style = 'suns_point_entry_%s' % level
            text_style = 'suns_point_text_%s' % level
            self.set_cell(ws, row, ss.ADDRESS_OFFSET_IDX + 1, values[ss.ADDRESS_OFFSET_IDX], entry_style)
            self.set_cell(ws, row, ss.GROUP_OFFSET_IDX + 1, values[ss.GROUP_OFFSET_IDX], entry_style)
            self.set_cell(ws, row, ss.NAME_IDX + 1, values[ss.NAME_IDX], entry_style)
            self.set_cell(ws, row, ss.VALUE_IDX + 1, values[ss.VALUE_IDX], entry_style)
            self.set_cell(ws, row, ss.COUNT_IDX + 1, values[ss.COUNT_IDX], entry_style)
            self.set_cell(ws, row, ss.TYPE_IDX + 1, values[ss.TYPE_IDX], entry_style)

            # don't put type size in xlsx unless point type is string
            if values[ss.TYPE_IDX] == 'string':
                self.set_cell(ws, row, ss.SIZE_IDX + 1, values[ss.SIZE_IDX], entry_style)
            else:
                self.set_cell(ws, row, ss.SIZE_IDX + 1, '', entry_style)

            self.set_cell(ws, row, ss.SCALE_FACTOR_IDX + 1, values[ss.SCALE_FACTOR_IDX], entry_style)
            self.set_cell(ws, row, ss.UNITS_IDX + 1, values[ss.UNITS_IDX], entry_style)
            self.set_cell(ws, row, ss.ACCESS_IDX + 1, values[ss.ACCESS_IDX], entry_style)
            self.set_cell(ws, row, ss.MANDATORY_IDX + 1, values[ss.MANDATORY_IDX], entry_style)
            self.set_cell(ws, row, ss.STATIC_IDX + 1, values[ss.STATIC_IDX], entry_style)
            self.set_info(ws, row, values, text_style)

        def set_symbol(self, ws, row, values):
            for i in range(len(values)):
                self.set_cell(ws, row, i + 1, '', 'suns_symbol_entry')
            self.set_cell(ws, row, ss.NAME_IDX + 1, values[ss.NAME_IDX])
            self.set_cell(ws, row, ss.VALUE_IDX + 1, values[ss.VALUE_IDX])
            self.set_info(ws, row, values, 'suns_symbol_text')

        def set_comment(self, ws, row, values):
            ws.merge_cells('A%s:%s%s' % (row, chr(65+len(values)-1), row))
            self.set_cell(ws, row, 1, values[0], 'suns_comment')

        def set_hdr(self, ws, values):
            for i in range(len(values)):
                self.set_cell(ws, 1, i + 1, values[i], 'suns_hdr')
                width = column_width[i]
                if width:
                    ws.column_dimensions[chr(65+i)].width = column_width[i]

        def spreadsheet_to_xlsx(self, mid, spreadsheet):
            if self.filename:
                raise ValueError('Workbooks opened with existing file are read only')
            has_notes = 'Notes' in spreadsheet[0]
            info = False
            label = None
            description = None
            notes = None
            level = 1

            ws = self.wb.create_sheet(title=str(mid))
            self.set_hdr(ws, spreadsheet[0])
            row = 2
            for values in spreadsheet[1:]:
                # point - has type
                etype = values[ss.TYPE_IDX]
                if etype:
                    # group
                    if etype in mdef.group_types:
                        level = len(values[ss.NAME_IDX].split('.'))
                        self.set_group(ws, row, values, level)
                        if not info:
                            label = values[ss.LABEL_IDX]
                            description = values[ss.DESCRIPTION_IDX]
                            if has_notes:
                                notes = values[ss.NOTES_IDX]
                            info = True
                    # point
                    elif etype in mdef.point_type_info:
                        self.set_point(ws, row, values, level)
                    else:
                        raise Exception('Unknown element type: %s' % etype)
                elif values[ss.NAME_IDX]:
                    # symbol - has name and value with no type
                    if values[ss.VALUE_IDX] is not None and values[ss.VALUE_IDX] != '':
                        self.set_symbol(ws, row, values)
                # comment - no name, value, or type
                elif values[0]:
                    self.set_comment(ws, row, values)
                row += 1

            if self.ws_models is not None:
                row = self.ws_models.max_row + 1
                self.set_cell(self.ws_models, row, 1, str(mid), 'suns_entry')
                cell = self.set_cell(self.ws_models, row, 2, label, 'suns_hyper')
                cell.hyperlink = '#%s!%s' % (str(mid), 'A1')
                self.set_cell(self.ws_models, row, 3, description, 'suns_text')

        def to_xlsx(self, model_def):
            mid = model_def[mdef.ID]
            spreadsheet = ss.to_spreadsheet(model_def)
            self.spreadsheet_to_xlsx(mid, spreadsheet)

        def create_error_sheet(self, mid, err_msg):
            ws = self.wb.create_sheet(title=str(mid))
            ws.column_dimensions['A'].width = 40
            ws['A1'] = 'Model Definition Errors'
            ws['A1'].font = styles.Font(bold=True)
            ws['A1'].alignment = styles.Alignment(horizontal='center')
            ws['A2'].alignment = styles.Alignment(horizontal='center', wrap_text=True)
            ws['A2'] = err_msg

except:
    # provide indication the openpyxl library not available
    class ModelWorkbook(object):
        def __init__(self, filename=None, model_dir=None, license_summary=False):
            raise Exception('openpyxl library not installed, it is required for working with .xlsx files')

if __name__ == "__main__":
    pass
