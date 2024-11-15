import sunspec2.xlsx as xlsx
import pytest
import openpyxl
import openpyxl.styles as styles
import json


def test___init__():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    assert wb.filename == './sunspec2/tests/test_data/wb_701-705.xlsx'
    assert wb.params == {}

    wb2 = xlsx.ModelWorkbook()
    assert wb2.filename is None
    assert wb2.params == {}


def test_get_models():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    assert wb.get_models() == [701, 702, 703, 704, 705]
    wb2 = xlsx.ModelWorkbook()
    assert wb2.get_models() == []


def test_save(tmp_path):
    wb = xlsx.ModelWorkbook()
    wb.save(tmp_path / 'test.xlsx')
    wb2 = xlsx.ModelWorkbook(filename=tmp_path / 'test.xlsx')
    iter_rows = wb2.xlsx_iter_rows(wb.wb['Index'])
    assert next(iter_rows) == ['Model', 'Label', 'Description']


def test_xlsx_iter_rows():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    iter_rows = wb.xlsx_iter_rows(wb.wb['704'])
    assert next(iter_rows) == ['Address Offset', 'Group Offset', 'Name',
                               'Value', 'Count', 'Type', 'Size', 'Scale Factor',
                               'Units', 'RW Access (RW)', 'Mandatory (M)', 'Static (S)',
                               'Label', 'Description', 'Detailed Description', 'Standards']
    assert next(iter_rows) == [None, None, 'DERCtlAC', None, None, 'group',
                               None, None, None, None, None, None, 'DER AC Controls',
                               'DER AC controls model.', None, None]


def test_spreadsheet_from_xlsx():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    assert wb.spreadsheet_from_xlsx(704)[0:2] == [['Address Offset', 'Group Offset', 'Name', 'Value', 'Count',
                                                   'Type', 'Size', 'Scale Factor', 'Units', 'RW Access (RW)',
                                                   'Mandatory (M)', 'Static (S)', 'Label', 'Description',
                                                   'Detailed Description', 'Standards'],
                                                  ['', '', 'DERCtlAC', None, None, 'group', None, None, None,
                                                   None, None, None, 'DER AC Controls', 'DER AC controls model.', None,
                                                   None]]


def sort_nested_dicts(d):
    for key, value in d.items():
        if isinstance(value, dict):
            d[key] = sort_nested_dicts(value)  # Sort nested dictionaries
        elif key == 'points' and isinstance(value, list):
            d[key] = sorted(value, key=lambda x: x['name'])
        elif isinstance(value, list) and len(value) > 1:
            d[key] = sorted(value, key=lambda x: sorted(x.items()) if isinstance(x, dict) else x)
    return dict(sorted(d.items()))


# need deep diff to compare from_xlsx to json file, right now just compares with its own output
def test_from_xlsx():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    with open('./sunspec2/models/json/model_704.json') as f:
        from_xlsx_output = json.load(f)

    a = sort_nested_dicts(wb.from_xlsx(704))
    b = sort_nested_dicts(from_xlsx_output)
    assert a == b


def test_set_cell():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    with pytest.raises(ValueError) as exc:
        wb.set_cell(wb.wb['704'], 1, 2, 3)
    assert 'Workbooks opened with existing file are read only' in str(exc.value)

    wb2 = xlsx.ModelWorkbook()
    assert wb2.set_cell(wb2.wb['Index'], 2, 1, 3, style='suns_comment').value == 3


def test_set_info():
    wb = xlsx.ModelWorkbook()
    values = [''] * 14
    values[13] = 'description'
    values[12] = 'label'
    wb.set_info(wb.wb['Index'], 2, values)
    iter_rows = wb.xlsx_iter_rows(wb.wb['Index'])
    next(iter_rows)
    assert next(iter_rows) == [None, None, None, None, None, None,
                               None, None, None, None, None, None, 'label', 'description']


def test_set_group():
    wb = xlsx.ModelWorkbook()
    values = [''] * 16
    values[2] = 'name'
    values[5] = 'type'
    values[4] = 'count'
    values[13] = 'description'
    values[12] = 'label'
    wb.set_group(wb.wb['Index'], 2, values, 2)
    iter_rows = wb.xlsx_iter_rows(wb.wb['Index'])
    next(iter_rows)
    assert next(iter_rows) == ['', '', 'name', '', 'count', 'type', '', '', '', '', '', '',
                               'label', 'description', '', '']


def test_set_point():
    wb = xlsx.ModelWorkbook()
    values = [''] * 16
    values[0] = 'addr_offset'
    values[1] = 'group_offset'
    values[2] = 'name'
    values[3] = 'value'
    values[4] = 'count'
    values[5] = 'type'
    values[6] = 'size'
    values[7] = 'sf'
    values[8] = 'units'
    values[9] = 'access'
    values[10] = 'mandatory'
    values[11] = 'static'
    wb.set_point(wb.wb['Index'], 2, values, 1)
    iter_rows = wb.xlsx_iter_rows(wb.wb['Index'])
    next(iter_rows)
    assert next(iter_rows) == ['addr_offset', 'group_offset', 'name', 'value', 'count', 'type', '',
                               'sf', 'units', 'access', 'mandatory', 'static', '', '', '', '']


def test_set_symbol():
    wb = xlsx.ModelWorkbook()
    values = [''] * 16
    values[2] = 'name'
    values[3] = 'value'
    values[12] = 'label'
    values[13] = 'description'
    wb.set_symbol(wb.wb['Index'], 2, values)
    iter_rows = wb.xlsx_iter_rows(wb.wb['Index'])
    next(iter_rows)  # skip header (['Model', 'Label', 'Description', None, ...])
    assert next(iter_rows) == ['', '', 'name', 'value', '', '', '',
                               '', '', '', '', '', 'label', 'description', '', '']


def test_set_comment():
    wb = xlsx.ModelWorkbook()
    wb.set_comment(wb.wb['Index'], 2, ['This is a comment'])
    iter_rows = wb.xlsx_iter_rows(wb.wb['Index'])
    next(iter_rows)
    assert next(iter_rows)[0] == 'This is a comment'


def test_set_hdr():
    wb = xlsx.ModelWorkbook()
    wb.set_hdr(wb.wb['Index'], ['This', 'is', 'a', 'test', 'header'])
    iter_rows = wb.xlsx_iter_rows(wb.wb['Index'])
    assert next(iter_rows) == ['This', 'is', 'a', 'test', 'header']


def test_spreadsheet_to_xlsx():
    wb = xlsx.ModelWorkbook(filename='./sunspec2/tests/test_data/wb_701-705.xlsx')
    with pytest.raises(ValueError) as exc:
        wb.spreadsheet_to_xlsx(702, [])
    assert 'Workbooks opened with existing file are read only' in str(exc.value)

    spreadsheet_smdx_304 = [
        ['Address Offset', 'Group Offset', 'Name', 'Value', 'Count', 'Type', 'Size', 'Scale Factor', 'Units',
         'RW Access (RW)', 'Mandatory (M)', 'Static (S)', 'Label', 'Description', 'Detailed Description', 'Standards'],
        ['', '', 'inclinometer', '', '', 'group', '', '', '', '', '', '', 'Inclinometer Model',
         'Include to support orientation measurements', '', ''],
        [0, '', 'ID', 304, '', 'uint16', '', '', '', '', 'M', 'S', 'Model ID', 'Model identifier', '', ''],
        [1, '', 'L', '', '', 'uint16', '', '', '', '', 'M', 'S', 'Model Length', 'Model length', '', ''],
        ['', '', 'inclinometer.incl', '', 0, 'group', '', '', '', '', '', '', '', '', '', ''],
        ['', 0, 'Inclx', '', '', 'int32', '', -2, 'Degrees', '', 'M', '', 'X', 'X-Axis inclination', '', ''],
        ['', 2, 'Incly', '', '', 'int32', '', -2, 'Degrees', '', '', '', 'Y', 'Y-Axis inclination', '', ''],
        ['', 4, 'Inclz', '', '', 'int32', '', -2, 'Degrees', '', '', '', 'Z', 'Z-Axis inclination', '', '']
    ]
    wb2 = xlsx.ModelWorkbook()
    wb2.spreadsheet_to_xlsx(304, spreadsheet_smdx_304)
    iter_rows = wb2.xlsx_iter_rows(wb2.wb['304'])
    for row in spreadsheet_smdx_304:
        assert next(iter_rows) == row


def test_to_xlsx(tmp_path):
    spreadsheet_smdx_304 = [
        ['Address Offset', 'Group Offset', 'Name', 'Value', 'Count', 'Type', 'Size', 'Scale Factor', 'Units',
         'RW Access (RW)', 'Mandatory (M)', 'Static (S)', 'Label', 'Description', 'Detailed Description', 'Standards'],
        ['', '', 'inclinometer', '', '', 'group', '', '', '', '', '', '', 'Inclinometer Model',
         'Include to support orientation measurements', '', ''],
        [0, '', 'ID', 304, '', 'uint16', '', '', '', '', 'M', 'S', 'Model ID', 'Model identifier', '', ''],
        [1, '', 'L', '', '', 'uint16', '', '', '', '', 'M', 'S', 'Model Length', 'Model length', '', ''],
        ['', '', 'inclinometer.incl', '', 0, 'group', '', '', '', '', '', '', '', '', '', ''],
        ['', 0, 'Inclx', '', '', 'int32', '', -2, 'Degrees', '', 'M', '', 'X', 'X-Axis inclination', '', ''],
        ['', 2, 'Incly', '', '', 'int32', '', -2, 'Degrees', '', '', '', 'Y', 'Y-Axis inclination', '', ''],
        ['', 4, 'Inclz', '', '', 'int32', '', -2, 'Degrees', '', '', '', 'Z', 'Z-Axis inclination', '', '']
    ]
    with open('./sunspec2/models/json/model_304.json') as f:
        m_703 = json.load(f)
    wb = xlsx.ModelWorkbook()
    wb.to_xlsx(m_703)
    iter_rows = wb.xlsx_iter_rows(wb.wb['304'])
    for row in spreadsheet_smdx_304:
        assert next(iter_rows) == row
